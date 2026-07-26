# -*- coding: utf-8 -*-
"""
Сборка Excel-системы контроля качества проб воды:
  - "Нормативы"      — справочник ПДК/нормативов по параметрам
  - "Градуировка_Fe" — коэффициенты пересчёта сигнала в концентрацию (МНК)
  - "Данные_ввод"    — сырой ввод измерений (то, что раньше писали в журнал)
  - "Результаты"     — автосвод: расчёт показателей, статус по каждому
                       параметру и итоговый статус пробы (формулами)
  - "Дашборд"        — сводка по месяцам (доля соответствия, число
                       превышений по параметрам) + встроенный график
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

ROOT = Path(__file__).resolve().parent.parent
RAW_CSV = ROOT / "data" / "raw_measurements.csv"
OUT_XLSX = ROOT / "excel" / "qc_workbook.xlsx"

df = pd.read_csv(RAW_CSV, encoding="utf-8-sig")

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=13, color="1F3864")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="808080")
BOLD = Font(name=FONT_NAME, bold=True)
BASE_FONT = Font(name=FONT_NAME)
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
FAIL_FILL = PatternFill("solid", fgColor="FCE4E4")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# =========================================================================
# 1. Лист "Нормативы" — справочник допустимых значений
# =========================================================================
ws_norm = wb.active
ws_norm.title = "Нормативы"

norm_headers = ["Параметр", "Мин.", "Макс.", "Единицы", "Источник / примечание"]
for j, h in enumerate(norm_headers, start=1):
    c = ws_norm.cell(row=1, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", wrap_text=True)

norm_rows = [
    ("Fe (концентрация)", None, 0.3, "мг/дм³", "СанПиН 1.2.3685-21, питьевая вода"),
    ("pH", 6.0, 9.0, "ед. pH", "СанПиН 1.2.3685-21, допустимый диапазон"),
    ("Электропроводность", None, 1000, "мкСм/см", "Внутренний технологический норматив лаборатории (не ПДК)"),
]
for i, row in enumerate(norm_rows, start=2):
    for j, val in enumerate(row, start=1):
        c = ws_norm.cell(row=i, column=j, value=val)
        c.font = BASE_FONT
        c.border = BORDER
        c.alignment = Alignment(horizontal="center" if j < 4 else "left", wrap_text=True)

for col, width in zip("ABCDE", [22, 10, 10, 12, 46]):
    ws_norm.column_dimensions[col].width = width

# именованные ссылки на нормативы для использования в формулах
FE_MAX = "Нормативы!$C$2"
PH_MIN = "Нормативы!$B$3"
PH_MAX = "Нормативы!$C$3"
EC_MAX = "Нормативы!$C$4"

# =========================================================================
# 2. Лист "Градуировка_Fe" — коэффициенты пересчёта (из кейса №1)
# =========================================================================
ws_cal = wb.create_sheet("Градуировка_Fe")
ws_cal.cell(row=1, column=1, value="Градуировка Fe (перенесена из кейса калибровки)").font = TITLE_FONT
ws_cal.merge_cells("A1:B1")
ws_cal.cell(row=3, column=1, value="k (тангенс угла наклона)").font = BOLD
ws_cal.cell(row=4, column=1, value="b (свободный член)").font = BOLD
k_cell = ws_cal.cell(row=3, column=2, value=0.8373)
b_cell = ws_cal.cell(row=4, column=2, value=0.0113)
for c in (k_cell, b_cell):
    c.font = BASE_FONT
    c.border = BORDER
    c.number_format = "0.0000"
note = ws_cal.cell(row=6, column=1,
                    value="Коэффициенты получены методом наименьших квадратов в отдельном "
                          "кейсе калибровки (см. repo metal-calibration-analytics). "
                          "Здесь используются для перевода сырого сигнала в концентрацию.")
note.font = NOTE_FONT
ws_cal.merge_cells("A6:F8")
note.alignment = Alignment(wrap_text=True, vertical="top")
ws_cal.column_dimensions["A"].width = 26
ws_cal.column_dimensions["B"].width = 12
K_REF = "Градуировка_Fe!$B$3"
B_REF = "Градуировка_Fe!$B$4"

# =========================================================================
# 3. Лист "Данные_ввод" — сырой ввод (то, что раньше писали в журнал)
# =========================================================================
ws_in = wb.create_sheet("Данные_ввод")
in_headers = ["Дата", "ID_пробы", "Точка_отбора", "A_Fe_сигнал",
              "pH_измерено", "Температура_C", "Электропроводность_мкСм_см"]
for j, h in enumerate(in_headers, start=1):
    c = ws_in.cell(row=1, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", wrap_text=True)

for i, row in enumerate(df.itertuples(index=False), start=2):
    for j, val in enumerate(row, start=1):
        c = ws_in.cell(row=i, column=j, value=val)
        c.font = BASE_FONT
        c.border = BORDER

for col, width in zip("ABCDEFG", [12, 10, 20, 12, 12, 13, 20]):
    ws_in.column_dimensions[col].width = width
ws_in.freeze_panes = "A2"
n_rows = ws_in.max_row

# =========================================================================
# 4. Лист "Результаты" — автоматический свод (расчёт + статус)
# =========================================================================
ws_res = wb.create_sheet("Результаты")
res_headers = [
    "Дата", "ID_пробы", "Точка_отбора",
    "Fe, мг/дм³", "Статус Fe",
    "pH", "Статус pH",
    "Электропроводность, мкСм/см", "Статус ЭП",
    "Итоговый статус пробы",
]
for j, h in enumerate(res_headers, start=1):
    c = ws_res.cell(row=1, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", wrap_text=True)

for i in range(2, n_rows + 1):
    r = i
    ws_res.cell(row=r, column=1, value=f"=Данные_ввод!A{r}").font = BASE_FONT
    ws_res.cell(row=r, column=2, value=f"=Данные_ввод!B{r}").font = BASE_FONT
    ws_res.cell(row=r, column=3, value=f"=Данные_ввод!C{r}").font = BASE_FONT

    fe_formula = f"=(Данные_ввод!D{r}-{B_REF})/{K_REF}"
    fe_cell = ws_res.cell(row=r, column=4, value=fe_formula)
    fe_cell.number_format = "0.000"

    fe_status = ws_res.cell(row=r, column=5, value=f'=IF(D{r}<={FE_MAX},"норма","превышение")')
    ph_cell = ws_res.cell(row=r, column=6, value=f"=Данные_ввод!E{r}")
    ph_cell.number_format = "0.00"
    ph_status = ws_res.cell(row=r, column=7, value=f'=IF(AND(F{r}>={PH_MIN},F{r}<={PH_MAX}),"норма","превышение")')
    ec_cell = ws_res.cell(row=r, column=8, value=f"=Данные_ввод!G{r}")
    ec_status = ws_res.cell(row=r, column=9, value=f'=IF(H{r}<={EC_MAX},"норма","превышение")')
    final_status = ws_res.cell(row=r, column=10,
                                value=f'=IF(AND(E{r}="норма",G{r}="норма",I{r}="норма"),"Проба соответствует","Проба не соответствует")')

    for c in (fe_cell, fe_status, ph_cell, ph_status, ec_cell, ec_status, final_status):
        c.font = BASE_FONT
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

for col, width in zip("ABCDEFGHIJ", [12, 10, 20, 12, 12, 8, 12, 14, 12, 22]):
    ws_res.column_dimensions[col].width = width
ws_res.freeze_panes = "A2"

# условное форматирование статусов (закраска "норма"/"превышение")
from openpyxl.formatting.rule import CellIsRule
for col_letter in ("E", "G", "I", "J"):
    rng = f"{col_letter}2:{col_letter}{n_rows}"
    ws_res.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"норма"'], fill=OK_FILL))
    ws_res.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"превышение"'], fill=FAIL_FILL))
ws_res.conditional_formatting.add(
    f"J2:J{n_rows}", CellIsRule(operator="equal", formula=['"Проба соответствует"'], fill=OK_FILL))
ws_res.conditional_formatting.add(
    f"J2:J{n_rows}", CellIsRule(operator="equal", formula=['"Проба не соответствует"'], fill=FAIL_FILL))

# =========================================================================
# 5. Лист "Дашборд" — помесячная сводка + график
# =========================================================================
ws_dash = wb.create_sheet("Дашборд")
ws_dash.sheet_view.showGridLines = False
ws_dash.cell(row=1, column=1, value="Дашборд контроля качества проб (помесячно)").font = TITLE_FONT
ws_dash.merge_cells("A1:F1")

months = sorted(set(d[:7] for d in df["Дата"]))  # YYYY-MM
dash_headers = ["Месяц", "Всего проб", "Доля соответствия, %", "Превышений Fe", "Превышений pH", "Превышений ЭП"]
for j, h in enumerate(dash_headers, start=3):
    c = ws_dash.cell(row=3, column=j - 2, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", wrap_text=True)

for i, month in enumerate(months, start=4):
    ws_dash.cell(row=i, column=1, value=month).font = BASE_FONT
    total_f = f'=COUNTIF(Результаты!$A$2:$A${n_rows},"{month}*")'
    total_c = ws_dash.cell(row=i, column=2, value=total_f)
    pass_f = (f'COUNTIFS(Результаты!$A$2:$A${n_rows},"{month}*",'
              f'Результаты!$J$2:$J${n_rows},"Проба соответствует")')
    share_f = f"=IFERROR(ROUND({pass_f}/B{i}*100,1),0)"
    share_c = ws_dash.cell(row=i, column=3, value=share_f)
    share_c.number_format = "0.0"

    fe_fail = f'=COUNTIFS(Результаты!$A$2:$A${n_rows},"{month}*",Результаты!$E$2:$E${n_rows},"превышение")'
    ph_fail = f'=COUNTIFS(Результаты!$A$2:$A${n_rows},"{month}*",Результаты!$G$2:$G${n_rows},"превышение")'
    ec_fail = f'=COUNTIFS(Результаты!$A$2:$A${n_rows},"{month}*",Результаты!$I$2:$I${n_rows},"превышение")'
    ws_dash.cell(row=i, column=4, value=fe_fail).font = BASE_FONT
    ws_dash.cell(row=i, column=5, value=ph_fail).font = BASE_FONT
    ws_dash.cell(row=i, column=6, value=ec_fail).font = BASE_FONT

    for col in range(1, 7):
        cell = ws_dash.cell(row=i, column=col)
        cell.font = BASE_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")

last_dash_row = 3 + len(months)
for col, width in zip("ABCDEF", [12, 12, 18, 14, 14, 14]):
    ws_dash.column_dimensions[col].width = width

# встроенный график: доля соответствия по месяцам
chart = LineChart()
chart.title = "Доля проб, соответствующих нормативам, по месяцам (%)"
chart.y_axis.title = "% соответствия"
chart.x_axis.title = "Месяц"
chart.style = 10
data_ref = Reference(ws_dash, min_col=3, min_row=3, max_row=last_dash_row)
cats_ref = Reference(ws_dash, min_col=1, min_row=4, max_row=last_dash_row)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.height = 8
chart.width = 18
ws_dash.add_chart(chart, "H3")

bar = BarChart()
bar.type = "col"
bar.grouping = "stacked"
bar.title = "Превышения по параметрам, по месяцам"
bar.y_axis.title = "Число проб с превышением"
data_ref2 = Reference(ws_dash, min_col=4, max_col=6, min_row=3, max_row=last_dash_row)
bar.add_data(data_ref2, titles_from_data=True)
bar.set_categories(cats_ref)
bar.height = 8
bar.width = 18
ws_dash.add_chart(bar, "H20")

note_dash = ws_dash.cell(
    row=last_dash_row + 2, column=1,
    value="Все значения в таблице — формулы (COUNTIF/COUNTIFS), автоматически "
          "пересчитываются при добавлении новых проб на листе «Данные_ввод».",
)
note_dash.font = NOTE_FONT
ws_dash.merge_cells(start_row=last_dash_row + 2, start_column=1, end_row=last_dash_row + 3, end_column=6)
note_dash.alignment = Alignment(wrap_text=True, vertical="top")

wb.save(OUT_XLSX)
print("Сохранено:", OUT_XLSX)
